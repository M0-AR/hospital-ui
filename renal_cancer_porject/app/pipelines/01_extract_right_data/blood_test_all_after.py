import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def parse_biochemistry_value(cell_content, biochemistry_keys):
    """
    Parses the value for a given biochemistry key from the cell content.

    Parameters:
        cell_content (str): The content of the cell to parse.
        biochemistry_keys (list): A list of biochemistry measurements to extract.

    Returns:
        tuple: A tuple containing the key and the parsed value, or (None, None) if no key is found.
    """
    for key in biochemistry_keys:
        if key.lower() in cell_content.lower():
            value_str = cell_content.split(';')[-1].split(':')[-1].strip()
            value_str = value_str.replace(',', '.').lstrip('<')
            try:
                return key, float(value_str)
            except ValueError:
                return key, value_str
    return None, None


def process_excel_file(file_path, biochemistry_keys, target_date_obj):
    """
    Processes the Excel file of a single patient to extract biochemistry data.

    Parameters:
        file_path (str): Full path to the patient's Excel file.
        biochemistry_keys (list): Biochemistry measurements to look for.
        target_date_obj (datetime): The cutoff datetime object for data extraction.

    Returns:
        dict: Dictionary with the latest biochemistry data and the closest date.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2} \d{2}:\d{2})')

    closest_date = None
    most_recent_values = {key: None for key in biochemistry_keys}
    closest_row_data = {}

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                cell_content = cell.strip().replace('_x000D_', '')
                date_match = date_pattern.match(cell_content)

                if date_match:
                    row_date_str = date_match.groups()[0]
                    row_date_obj = datetime.strptime(row_date_str, "%d-%m-%y %H:%M")
                    # if row_date_obj > target_date_obj:
                    #     return closest_row_data, closest_date
                    if row_date_obj < target_date_obj:
                        continue
                    if closest_date is None or row_date_obj > closest_date:
                        closest_date = row_date_obj
                        closest_row_data = most_recent_values.copy()

                elif closest_date:
                    key, value = parse_biochemistry_value(cell_content, biochemistry_keys)
                    if key:
                        closest_row_data[key] = value
                        most_recent_values[key] = value

    return closest_row_data, closest_date


def extract_latest_patient_biochemistry_data_before_operation(file_name, base_directory, operation_date,
                                                              biochemistry_keys):
    """
    Orchestrates the extraction of the latest biochemistry data for each patient before an operation date.

    Parameters:
        file_name (str): The name of the Excel file within each patient's subdirectory.
        base_directory (str): Path to the directory containing patient subdirectories.
        operation_date (str): The cutoff date and time as a string in the format 'dd-mm-yy HH:MM'.
        biochemistry_keys (list): List of biochemistry measurements to extract.

    Returns:
        pd.DataFrame: DataFrame containing the latest biochemistry data for each patient.
    """
    operation_date_obj = datetime.strptime(operation_date, "%d-%m-%y %H:%M")
    patient_data = []

    for patient_dir in os.listdir(base_directory):
        patient_path = os.path.join(base_directory, patient_dir)
        if os.path.isdir(patient_path):
            file_path = os.path.join(patient_path, file_name)
            if os.path.isfile(file_path):
                closest_row_data, closest_date = process_excel_file(file_path, biochemistry_keys, operation_date_obj)
                if closest_date:
                    patient_data.append({
                        'cpr': patient_dir,
                        'blood_date': closest_date.strftime("%d-%m-%y %H:%M"),
                        **closest_row_data
                    })

    return pd.DataFrame(patient_data)


# # Example usage
# file_name = 'blood_test.xlsx'
# base_directory = 'C:\\src\\hospital-ui\\renal_cancer_porject\\data'
# operation_date = '01-01-19 08:00'  # Operation date in 'dd-mm-yy HH:MM' format
# biochemistry_keys = ['crp', 'Leukocytter', 'Neutrofili', 'Sedimentationsrate', 'Hæmoglobin']
#
# blood_test_data = extract_latest_patient_biochemistry_data_before_operation(file_name, base_directory, operation_date,
#                                                                             biochemistry_keys)
# print(blood_test_data)

import pandas as pd
from datetime import datetime
import os

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re

from openpyxl import load_workbook
from datetime import datetime
import re


# def parse_biochemistry_value(cell_content):
#     """
#     Tries to parse any potential biochemistry value from the cell content.
#     Assumes format "Key: Value" within the cell.
#
#     Parameters:
#         cell_content (str): The content of the cell to parse.
#
#     Returns:
#         tuple: A tuple containing the key and the parsed value, or (None, None) if no key:value format is found.
#     """
#     if ':' in cell_content:
#         parts = cell_content.split(':')
#         if len(parts) > 1:
#             key = parts[0].strip()
#             value_str = parts[1].strip().replace(',', '.').lstrip('<')  # Normalize decimal and remove less than signs
#             try:
#                 value = float(value_str)  # Convert to float if possible
#             except ValueError:
#                 value = value_str  # Keep as string if conversion fails
#             return key, value
#     return None, None
#
#
# def process_excel_file(file_path, target_date_obj):
#     """
#     Processes the Excel file of a single patient to extract biochemistry data dynamically.
#
#     Parameters:
#         file_path (str): Full path to the patient's Excel file.
#         target_date_obj (datetime): The cutoff datetime object for data extraction.
#
#     Returns:
#         dict: Dictionary with the latest biochemistry data and the closest date.
#     """
#     workbook = load_workbook(file_path)
#     sheet = workbook.active
#     date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2} \d{2}:\d{2})')
#
#     closest_date = None
#     closest_row_data = {}
#
#     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
#         for cell in row:
#             if cell and isinstance(cell, str):
#                 cell_content = cell.strip().replace('_x000D_', '')
#                 date_match = date_pattern.search(cell_content)
#                 if date_match:
#                     row_date_str = date_match.group(0)
#                     row_date = datetime.strptime(row_date_str, "%d-%m-%y %H:%M")
#                     if row_date > target_date_obj:
#                         break  # Stop processing if the date exceeds the target date
#                     if not closest_date or row_date > closest_date:
#                         closest_date = row_date
#                         closest_row_data = {}  # Reset closest data
#                 elif closest_date:
#                     key, value = parse_biochemistry_value(cell_content)
#                     if key:
#                         closest_row_data[key] = value  # Update with new value
#
#     return closest_row_data, closest_date


def extract_key_and_value(cell_content):
    """
    Extracts the key and value from a given string formatted as 'Key: Value'.

    Parameters:
        cell_content (str): The content of the cell in the format 'Key: Value'.

    Returns:
        tuple: A tuple containing the key and the parsed value.
    """
    # Split the content into key and value parts
    parts = cell_content.split(':')
    if len(parts) == 2:
        key = parts[0].strip()
        value_str = parts[1].strip()
        value_str = value_str.replace(',', '.').lstrip('<')

        # Attempt to convert value to float, if possible
        try:
            value = float(value_str)
        except ValueError:
            value = value_str  # Keep as string if it's not a valid float

        return key, value

    return None, None  # Return None if the format isn't as expected

from datetime import timedelta
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

def process_excel_file(file_path, target_date_obj, days_after, max_values_per_key=5):
    """
    Processes the Excel file of a single patient to extract biochemistry data after a specific date plus additional days,
    without resetting collected data unless it surpasses the set limit of values per key.

    Parameters:
        file_path (str): Full path to the patient's Excel file.
        target_date_obj (datetime): The base datetime object for data extraction.
        days_after (int): Number of days after the target date to start data extraction.
        max_values_per_key (int): Maximum number of values to store for each key.

    Returns:
        dict: Dictionary with keys as biochemistry markers and values as lists of (date, value) tuples.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2} \d{2}:\d{2})')

    end_date = target_date_obj + timedelta(days=days_after)
    latest_data = {}
    latest_date = None

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                cell_content = cell.strip().replace('_x000D_', '')
                date_match = date_pattern.match(cell_content)

                if date_match:
                    row_date_str = date_match.groups()[0]
                    row_date_obj = datetime.strptime(row_date_str, "%d-%m-%y %H:%M")
                    if row_date_obj > end_date:  # Only consider dates after the end_date
                        if latest_date is None or row_date_obj > latest_date:
                            latest_date = row_date_obj  # Update the latest date
                            # Do not reset latest_data here
                elif latest_date:  # Only process cells if a valid date has been set and is after the end_date
                    key, value = extract_key_and_value(cell_content)
                    if key:
                        if key not in latest_data:
                            latest_data[key] = []  # Initialize a new list for this key
                        # Append only up to the maximum number of allowed entries
                        if len(latest_data[key]) < max_values_per_key:
                            # latest_data[key].append((latest_date, value))
                            latest_data[key].append((latest_date.strftime('%Y-%m-%d %H:%M:%S'), value))

    return latest_data, latest_date

def read_patient_operation_dates(file_path):
    """
    Reads patient operation dates from an Excel file.

    Parameters:
        file_path (str): Path to the Excel file containing the operation dates.

    Returns:
        dict: Dictionary with patient ID as keys and operation date as Timestamp values.
    """
    df = pd.read_excel(file_path, parse_dates=['EarliestDiagnosisDate'])  # Ensure the dates are parsed
    dates = {row['PatientID']: row['EarliestDiagnosisDate'] for index, row in df.iterrows()}
    return dates



def extract_patient_biochemistry_data_after_operation(file_name, base_directory, patient_date_file, days_after):
    """
    Extracts the biochemistry data for each patient for a specified number of days after their operation date.

    Parameters:
        file_name (str): The name of the Excel file within each patient's subdirectory.
        base_directory (str): Path to the directory containing patient subdirectories.
        patient_date_file (str): Excel file with patient IDs and their specific operation dates.
        days_after (int): Number of days after the operation date to consider for data extraction.

    Returns:
        pd.DataFrame: DataFrame containing the latest biochemistry data for each patient within the specified period.
    """
    patient_dates = read_patient_operation_dates(patient_date_file)
    patient_data = []

    for patient_dir in os.listdir(base_directory):
        patient_path = os.path.join(base_directory, patient_dir)
        if os.path.isdir(patient_path):
            file_path = os.path.join(patient_path, file_name)
            if os.path.isfile(file_path) and patient_dir in patient_dates:
                operation_date_obj = patient_dates[patient_dir]  # This is already a datetime object
                latest_data, latest_date = process_excel_file(file_path, operation_date_obj, days_after)
                # print(latest_data)
                # print(latest_date)
                if latest_date:
                    patient_data.append({
                        'PatientID': patient_dir,
                        **latest_data
                    })

    return pd.DataFrame([{
        key if key == 'PatientID' else f'after_surgery_1mo_first5_{key}': value
        for key, value in data.items()
    } for data in patient_data])

# # Define the parameters
# file_name = 'blood_test.xlsx'
# base_directory = 'C:\\src\\hospital-ui\\renal_cancer_porject\\data'
# patient_date_file = 'C:\\Users\\md\\Downloads\\rcc.xlsx'
#
# # Call the function
# blood_test_data = extract_patient_biochemistry_data_after_operation(
#     file_name=file_name,
#     base_directory=base_directory,
#     patient_date_file=patient_date_file,
#     days_after=31
# )
#
# # Print or process the DataFrame
# print(blood_test_data)
