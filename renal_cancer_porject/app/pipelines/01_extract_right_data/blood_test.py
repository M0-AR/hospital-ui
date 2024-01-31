import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def parse_biochemistry_value(cell_content, biochemistry_keys):
    """
    Parses the value for a given biochemistry key from the cell content.

    Parameters:
        cell_content (str): The content of the cell to parse.
        biochemistry_keys (list): A list of biochemistry measurements to extract.

    Returns:
        tuple: A tuple containing the key and the parsed value, or (None, None) if no key is found.
    """
    for key in biochemistry_keys:
        if key.lower() in cell_content.lower():
            value_str = cell_content.split(';')[-1].split(':')[-1].strip()
            value_str = value_str.replace(',', '.').lstrip('<')
            try:
                return key, float(value_str)
            except ValueError:
                return key, value_str
    return None, None


def process_excel_file(file_path, biochemistry_keys, target_date_obj):
    """
    Processes the Excel file of a single patient to extract biochemistry data.

    Parameters:
        file_path (str): Full path to the patient's Excel file.
        biochemistry_keys (list): Biochemistry measurements to look for.
        target_date_obj (datetime): The cutoff datetime object for data extraction.

    Returns:
        dict: Dictionary with the latest biochemistry data and the closest date.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2} \d{2}:\d{2})')

    closest_date = None
    most_recent_values = {key: None for key in biochemistry_keys}
    closest_row_data = {}

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                cell_content = cell.strip().replace('_x000D_', '')
                date_match = date_pattern.match(cell_content)

                if date_match:
                    row_date_str = date_match.groups()[0]
                    row_date_obj = datetime.strptime(row_date_str, "%d-%m-%y %H:%M")
                    if row_date_obj > target_date_obj:
                        return closest_row_data, closest_date
                    if closest_date is None or row_date_obj > closest_date:
                        closest_date = row_date_obj
                        closest_row_data = most_recent_values.copy()

                elif closest_date:
                    key, value = parse_biochemistry_value(cell_content, biochemistry_keys)
                    if key:
                        closest_row_data[key] = value
                        most_recent_values[key] = value

    return closest_row_data, closest_date


def extract_latest_patient_biochemistry_data_before_operation(file_name, base_directory, operation_date,
                                                              biochemistry_keys):
    """
    Orchestrates the extraction of the latest biochemistry data for each patient before an operation date.

    Parameters:
        file_name (str): The name of the Excel file within each patient's subdirectory.
        base_directory (str): Path to the directory containing patient subdirectories.
        operation_date (str): The cutoff date and time as a string in the format 'dd-mm-yy HH:MM'.
        biochemistry_keys (list): List of biochemistry measurements to extract.

    Returns:
        pd.DataFrame: DataFrame containing the latest biochemistry data for each patient.
    """
    operation_date_obj = datetime.strptime(operation_date, "%d-%m-%y %H:%M")
    patient_data = []

    for patient_dir in os.listdir(base_directory):
        patient_path = os.path.join(base_directory, patient_dir)
        if os.path.isdir(patient_path):
            file_path = os.path.join(patient_path, file_name)
            if os.path.isfile(file_path):
                closest_row_data, closest_date = process_excel_file(file_path, biochemistry_keys, operation_date_obj)
                if closest_date:
                    patient_data.append({
                        'cpr': patient_dir,
                        'blood_date': closest_date.strftime("%d-%m-%y %H:%M"),
                        **closest_row_data
                    })

    return pd.DataFrame(patient_data)


# Example usage
file_name = 'blood_test.xlsx'
base_directory = 'C:\\src\\hospital-ui\\renal_cancer_porject\\data'
operation_date = '01-01-19 08:00'  # Operation date in 'dd-mm-yy HH:MM' format
biochemistry_keys = ['crp', 'Leukocytter', 'Neutrofili', 'Sedimentationsrate', 'Hæmoglobin']

blood_test_data = extract_latest_patient_biochemistry_data_before_operation(file_name, base_directory, operation_date,
                                                                            biochemistry_keys)
print(blood_test_data)

